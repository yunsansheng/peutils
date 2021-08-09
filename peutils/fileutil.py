# -*- coding: UTF-8 -*-

'''
Author: Henry Wang
Date: 2021-03-05 13:23
Short Description:

Change History:

'''
import os
from functools import partial
from pathlib import Path

'''
# Henry Wang, 2020-08-14 11:07 Changed:使用pathlib替换os.path 兼容windows
'''


def list_current_file(path='.', type='all', suffix='', not_prefix=(('~', '.'))):
    '''
    列出当前目录下的文件或者文件夹
    :param path: 默认当前目录
    :param type: 可选 file,folder,all 默认all
    :param suffix: 对文件夹和文件后缀过滤
    :param not_prefix: 对文件夹和文件前缀过滤，默认不要隐藏文件和临时文件
    :return:文件或文件夹的集合
    '''
    p = Path(path)

    if type == 'all':
        return [x.resolve().as_posix() for x in p.iterdir()
                if x.name.endswith(suffix) and not x.name.startswith(not_prefix)]
    elif type == "file":
        return [x.resolve().as_posix() for x in p.iterdir() if x.is_file() and  x.name.endswith(suffix) and not x.name.startswith(not_prefix)]
    elif type == "folder":
        return [x.resolve().as_posix() for x in p.iterdir() if x.is_dir() and  x.name.endswith(suffix) and not x.name.startswith(not_prefix)]
    else:
        raise Exception(f"type: {type} not defined.")


def list_files_deep(path='.', suffix='', not_prefix=(('~', '.'))):
    '''
    :param path: 默认当前目录 '.'
    :param suffix: 文件后缀，单个或者元组
    :param not_prefix: 单个或者元组,默认去掉隐藏文件和临时文件
    :return: 文件全路径集合
    '''
    files = []
    all_files = list(Path(path).glob('**/*.*'))  # 过滤出来的是文件

    # # Henry Wang, 2021-01-28 14:01 Changed:将suffix转成小写
    if isinstance(suffix,str)==True:
        suffix = suffix.lower()
    elif isinstance(suffix,tuple) ==True:
        suffix = tuple([x.lower() for x in suffix])

    # Henry Wang, 2021-05-25 14:18 Changed:增加os.path.isfile的判断
    for filpath in all_files:
        if filpath.is_file() ==True and filpath.name.lower().endswith(suffix) and not filpath.name.startswith(not_prefix):
            files.append(filpath.resolve().as_posix())

    return files


### 替换生成新地址，并创建文件夹
### 老的地址必须包含在
def replace_path_and_create(filepath, old_dir, new_dir, add_path: list):
    old_dir = Path(old_dir)
    new_dir = Path(new_dir)
    filepath = Path(filepath)

    if old_dir.resolve().as_posix() in filepath.resolve().as_posix():
        new_fl = filepath.resolve().as_posix().replace(old_dir.resolve().as_posix(), new_dir.resolve().as_posix())
        if add_path != None:
            p = Path(new_fl).resolve()
            new_fl = p.parent.joinpath(*add_path, p.name).resolve().as_posix()
        return new_fl
    else:
        raise Exception(f"{old_dir} not in {new_dir}")


'''
自定义方法，从A目录到B目录，找到对应后缀文件，循环并生成新的文件
'''


class Hanlde_Move_File():
    def __init__(self, origin_path, out_path, process_func, in_suffix, out_suffix, add_path=None):
        self.origin_path = origin_path
        self.out_path = out_path
        self.change_path = partial(replace_path_and_create, old_dir=origin_path, new_dir=out_path,
                                   add_path=add_path)  # 只要传filepath
        self.process_func = process_func
        self.in_suffix = in_suffix.strip()
        self.out_suffix = out_suffix.strip()
        self.add_path = add_path

    def main(self):
        print(f'handle {self.process_func.__name__.upper()} from: {self.origin_path} to: {self.out_path}')
        files = list_files_deep(self.origin_path, suffix=self.in_suffix)
        print(f'处理文件总数: {len(files)}')
        out_pre_file = list_files_deep(self.out_path, suffix=self.out_suffix)
        if len(out_pre_file) != 0:
            print(f'Warining: 目标文件夹，该类型文件存在 {out_pre_file} 个历史文件，请确认操作，无问题可忽略本提示')

        for filename in files:
            try:
                out_change_path = self.change_path(filename)[:-len(self.in_suffix)] + self.out_suffix
                out_path_parent = Path(out_change_path).resolve().parent
                out_path_parent.mkdir(parents=True, exist_ok=True)  # 递归创建目录，如果已存在不报错
                self.process_func(filename, out_change_path)

            except Exception as e:
                print(f'filename {filename} Error')

        out_file = list_files_deep(self.out_path, suffix=self.out_suffix)
        print(f"输出文件总数: {len(out_file)}")


'''
文件名相同的情况下检查文件对齐,不同后缀文件数量对齐,可以在相同文件夹，也可以在不同文件夹
'''



import csv


def getCsvData(csvPath):
    csv.field_size_limit(500 * 1024 * 1024)
    with open(csvPath, encoding='utf-8-sig') as f:
        data = csv.reader(f)
        header = next(data)
        data = list(data)
    return data, header


def saveCsvData(savePath, data: list, header=None, newline='\n'):
    csv.field_size_limit(500 * 1024 * 1024)
    with open(savePath, 'w', encoding='utf-8-sig', newline=newline) as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL, lineterminator="\n")
        if header:
            writer.writerow(header)
        writer.writerows(data)

def getExcelData(excelPath,sheetname=None):
    from openpyxl import load_workbook
    wb = load_workbook(excelPath)
    if sheetname is None:
        ws  = wb.active
    else:
        ws = wb[sheetname]
    all_data =ws.rows
    header = [x.value for x in next(all_data)]
    data = [[x.value for x in row] for row in all_data]
    return data,header


def saveExcelData(savePath,data: list, header=None):
    from openpyxl import Workbook
    wb = Workbook()
    ws=wb.active
    if header is not None:
        ws.append(header)
    for row in data:
        ws.append(row)
    wb.save(savePath)
